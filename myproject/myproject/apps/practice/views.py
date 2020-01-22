from django.http import HttpResponse
from django.shortcuts import render
from .models import Student

import os
from openpyxl import load_workbook



def index(request):
    students = []
    for i in Student.objects.all():
        students.append([i.student_Number, i.student_Group, i.student_Name, i.student_Enterprise, i.student_Contract,
                         i.student_Points, i.student_Rating, i.student_The_theme_of_work,
                         i.student_The_topic_of_work_in_English, i.student_Last_name_First_name_in_English,
                         i.student_Manager, i.student_Email])

    return render(request, 'practice/list.html', {"students": students})
    #return HttpResponse(Student.objects.all())

def student(request):

    module_dir = os.path.dirname(__file__)
    excel_file = os.path.join(module_dir, 'resources/excel/practice.xlsx')

    wb2 = load_workbook(excel_file, data_only=True)
    # print(wb2.sheetnames)
    # print(wb2['Переддипломна 2019'])
    ws = wb2["Page1"]
    rows = ws.rows
    next(rows)

    stud = []

    for row in rows:
        if row[1].value != None:
            #print('a ', row[1].value)
            tmp = []
            for i in range(13):
                tmp.append(row[i].value)
            stud.append(tmp)
            for cell in row:
                a = cell.value
                #print('b', a)

    """for i in range(len(stud)):
        #print(stud[i])
        a = Student(student_Specialty=stud[i][0], student_Group=stud[i][1], student_Number=stud[i][2],
                    student_Name=stud[i][3], student_Enterprise=stud[i][4],
                    student_Contract=stud[i][5], student_Points=stud[i][6], student_Rating=stud[i][7],
                    student_The_theme_of_work=stud[i][8],
                    student_The_topic_of_work_in_English=stud[i][9],
                    student_Last_name_First_name_in_English=stud[i][10],
                    student_Manager=stud[i][11],
                    student_Email=stud[i][12])

        a.save()
        
        """

    return HttpResponse("Student")


def student_data(request):
    return render(request, 'practice/student_data.html')

def student_add(request):

    if request.method == 'GET':
        if 'student_Name' in request.GET:
            return HttpResponse(request.GET['student_Name'])
        else:
            return HttpResponse("Hello")

