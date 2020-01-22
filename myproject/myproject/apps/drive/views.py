from django.http import HttpResponse
from django.shortcuts import render

from google.oauth2 import service_account
from googleapiclient.discovery import build
import os

from googleapiclient.http import MediaIoBaseDownload,MediaFileUpload
import io

import openpyxl

module_dir = os.path.dirname(__file__)  # get current directory
file_path = os.path.join(module_dir, 'resources/key/Students practice-fee9dbb84d6c.json')
SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = file_path
credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('drive', 'v3', credentials=credentials)

def index(request):

    download_file()

    excel_file = os.path.join(module_dir, 'resources/excel/practice.xlsx')
            #request.FILES["excel_file"]

        # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file, data_only=True)

        # getting all sheets
    sheets = wb.sheetnames
    print(sheets)

        # getting a particular sheet
    worksheet = wb["Page1"]
    print(worksheet)

        # getting active sheet
    active_sheet = wb.active
    print(active_sheet)

        # reading a cell
    print(worksheet["A1"].value)

    excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
    for row in worksheet.iter_rows():
        if row[1].value != None:
            tmp = []
            for i in range(13):
                if row[i].value == None:
                    tmp.append('')
                else:
                    tmp.append(row[i].value)
            #stud.append(tmp)
          #  row_data = list()
         #   for cell in row:
          #      if cell.value == None:
          ##          row_data.append(str(""))
           #     else:
          #          row_data.append(str(cell.value))
                    #print(cell.value)
            excel_data.append(tmp)

    return render(request, 'drive/index.html', {"excel_data": excel_data})



def show(request):

    # results = service.files().list(pageSize=10, fields="nextPageToken, files(id, name, mimeType, parents, "
    #                                                 "createdTime, permissions, "
    #                                                "quotaBytesUsed)", q="'1VqOlJQV1XTlIxSWTgh27VMEmG6JvSrMa' in parents"
    #                                                                  " and name contains 'Переддипломна'").execute()
    results = service.files().list(fields="files(id, name, createdTime)").execute()
    res = results['files']
    res2 = res[7:]
    return render(request, 'drive/list.html', {'S':res2})



def download_file():

    file_id = '1cZcRJGwW79rmXRoCc3GeuNtZfOqPCv19'
    # Ordinary download
    request = service.files().get_media(fileId=file_id)
    filename = os.path.join(module_dir, 'resources/excel/practice.xlsx')
    fh = io.FileIO(filename, 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        print("Download %d%%." % int(status.progress() * 100))